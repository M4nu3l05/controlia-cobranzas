from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SHEET_FREEZE = "A2"
DEFAULT_DATE_FORMAT = "DD/MM/YYYY"
DEFAULT_CURRENCY_FORMAT = '"$"#,##0'


def _looks_like_date_column(name: str) -> bool:
    txt = str(name or "").strip().lower()
    return "fecha" in txt


def _looks_like_currency_column(name: str) -> bool:
    txt = str(name or "").strip().lower()
    keys = ("monto", "saldo", "copago", "pago", "total pagos")
    return any(k in txt for k in keys)


def _parse_currency_value(value):
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    txt = str(value).strip().replace("$", "").replace(" ", "")
    if not txt:
        return None

    # Evita tratar fechas como moneda (ej: 17-11-2025).
    if re.fullmatch(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", txt):
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}( \d{2}:\d{2}(:\d{2})?)?", txt):
        return None

    try:
        if re.fullmatch(r"-?\d+(\.\d+)?", txt):
            return float(txt)

        if "," in txt and "." in txt:
            if txt.rfind(",") > txt.rfind("."):
                txt = txt.replace(".", "").replace(",", ".")
            else:
                txt = txt.replace(",", "")
            return float(txt)

        if txt.count(".") > 1:
            return float(txt.replace(".", ""))

        if txt.count(".") == 1:
            left, right = txt.split(".", 1)
            if len(right) == 3 and left.replace("-", "").isdigit():
                return float(f"{left}{right}")
            return float(txt)

        if txt.count(",") == 1:
            left, right = txt.split(",", 1)
            if len(right) == 3 and left.replace("-", "").isdigit():
                return float(f"{left}{right}")
            return float(txt.replace(",", "."))

        return float(txt)
    except Exception:
        # Si no se puede convertir, deja el valor como vacio numerico.
        return None


def _parse_date_value(value):
    if value is None or value == "":
        return pd.NaT

    if isinstance(value, bool):
        return pd.NaT

    txt = str(value).strip()
    if not txt:
        return pd.NaT

    # Formatos explicitos para evitar warnings y conversiones ambiguas.
    for fmt in (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return pd.to_datetime(txt, format=fmt, errors="raise")
        except Exception:
            pass

    # Fallback tolerante sin dayfirst forzado.
    return pd.to_datetime(txt, errors="coerce")


def _normalize_excel_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy() if df is not None else pd.DataFrame()

    for col_name in out.columns:
        if _looks_like_date_column(col_name):
            out[col_name] = out[col_name].apply(_parse_date_value)
        elif _looks_like_currency_column(col_name):
            out[col_name] = out[col_name].apply(_parse_currency_value)

    return out


def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_idx = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 40)


def _style_worksheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = SHEET_FREEZE
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for idx, col_name in enumerate(df.columns, start=1):
        if _looks_like_date_column(col_name):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = DEFAULT_DATE_FORMAT
        elif _looks_like_currency_column(col_name):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = DEFAULT_CURRENCY_FORMAT

    _autosize_worksheet(ws)


def write_excel_report(path: str, sheets: dict[str, pd.DataFrame]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_df = _normalize_excel_df(df)
            safe_df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_worksheet(writer.book[sheet_name], safe_df)
